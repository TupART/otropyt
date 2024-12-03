from flask import Flask, render_template, request, send_file
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
import os
import tempfile
from datetime import datetime

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Subir archivo
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file:
            filename = secure_filename(file.filename)
            
            # Guardar el archivo en un archivo temporal
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                file_path = temp_file.name
                file.save(file_path)
            
            # Leer el archivo con pandas
            df = pd.read_excel(file_path, header=1)  # Nombres de columnas en la fila 2 (índice 1)

            # Obtener solo los últimos 25 registros
            global last_data  # Guardar la referencia a los últimos 25 registros globalmente
            last_data = df[['Name', 'Surname', 'E-mail', 'Market', 'Va a ser PCC?', 'B2E User Name']].tail(25)
            data = last_data.to_dict(orient='records')
            
            return render_template('index.html', data=data)
    
    return render_template('index.html', data=None)

@app.route('/process', methods=['POST'])
def process():
    selected_rows = request.form.getlist('rows')  # Obtener filas seleccionadas
    selected_indices = [int(row) for row in selected_rows]  # Convertir a enteros

    # Validar que los índices estén dentro del rango de los últimos 25 registros
    if any(idx < 0 or idx >= len(last_data) for idx in selected_indices):
        return "Índice seleccionado está fuera de rango.", 400

    # Cargar la plantilla 'PlantillaSTEP4.xlsx'
    plantilla = 'PlantillaSTEP4.xlsx'
    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    # Contador para la fila de destino
    destination_row = 7  # Comenzar desde la fila 7 en la plantilla

    # Procesar las filas seleccionadas y rellenar la plantilla
    for idx in selected_indices:
        # Obtener los datos de last_data usando el índice
        name = last_data.iloc[idx]['Name']
        surname = last_data.iloc[idx]['Surname']
        email = last_data.iloc[idx]['E-mail']
        market = last_data.iloc[idx]['Market']
        pcc_status = last_data.iloc[idx]['Va a ser PCC?']
        b2e_username = last_data.iloc[idx]['B2E User Name']

        # Rellenar las columnas C, D y E
        ws[f'C{destination_row}'] = name
        ws[f'D{destination_row}'] = surname
        ws[f'E{destination_row}'] = email

        # Rellenar columna F (Primary phone)
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'F{destination_row}'] = "+4940210918145"
            elif market == 'France':
                ws[f'F{destination_row}'] = "+33180037979"
            elif market == 'Spain':
                ws[f'F{destination_row}'] = "+34932952130"
            elif market == 'Italy':
                ws[f'F{destination_row}'] = "+390109997099"
        else:
            ws[f'F{destination_row}'] = ""

        # Rellenar columna G (Workgroup)
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'G{destination_row}'] = "D_PCC"
            elif market == 'France':
                ws[f'G{destination_row}'] = "F_PCC"
            elif market == 'Spain':
                ws[f'G{destination_row}'] = "E_PCC"
            elif market == 'Italy':
                ws[f'G{destination_row}'] = "I_PCC"
        elif pcc_status == 'N':
            if market == 'DACH':
                ws[f'G{destination_row}'] = "D_Outbound"
            elif market == 'France':
                ws[f'G{destination_row}'] = "F_Outbound"
            elif market == 'Spain':
                ws[f'G{destination_row}'] = "E_Outbound"
        elif pcc_status == 'TL':
            if market == 'DACH':
                ws[f'G{destination_row}'] = "D_PCC"
            elif market == 'France':
                ws[f'G{destination_row}'] = "F_PCC"
            elif market == 'Spain':
                ws[f'G{destination_row}'] = "E_PCC"
            elif market == 'Italy':
                ws[f'G{destination_row}'] = "I_PCC"
        elif pcc_status == 'DS':
            if market == 'DACH':
                ws[f'G{destination_row}'] = "D_Outbound"
            elif market == 'France':
                ws[f'G{destination_row}'] = "F_Outbound"
            elif market == 'Spain':
                ws[f'G{destination_row}'] = "E_Outbound"

        # Rellenar columna H (Workgroup WA)
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'H{destination_row}'] = "D_WAPCC"
            elif market == 'France':
                ws[f'H{destination_row}'] = "F_WAPCC"
            elif market == 'Spain':
                ws[f'H{destination_row}'] = "E_WAPCC"
            elif market == 'Italy':
                ws[f'H{destination_row}'] = "I_WAPCC"
        elif pcc_status == 'DS':
            if market == 'DACH':
                ws[f'H{destination_row}'] = "D_WAPCC"
            elif market == 'France':
                ws[f'H{destination_row}'] = "F_WAPCC"
            elif market == 'Spain':
                ws[f'H{destination_row}'] = "E_WAPCC"
            elif market == 'Italy':
                ws[f'H{destination_row}'] = "I_WAPCC"

        # Rellenar columna I (Team)
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'I{destination_row}'] = "Team_D_CCH_PCC_1"
            elif market == 'France':
                ws[f'I{destination_row}'] = "Team_F_CCH_PCC_1"
            elif market == 'Spain':
                ws[f'I{destination_row}'] = "Team_E_CCH_PCC_1"
            elif market == 'Italy':
                ws[f'I{destination_row}'] = "Team_I_CCH_PCC_1"
        elif pcc_status == 'N':
            if market == 'DACH':
                ws[f'I{destination_row}'] = "Team_D_CCH_B2C_1"
            elif market == 'France':
                ws[f'I{destination_row}'] = "Team_F_CCH_B2C_1"
            elif market == 'Spain':
                ws[f'I{destination_row}'] = "Team_E_CCH_B2C_1"
        elif pcc_status == 'TL':
            if market == 'DACH':
                ws[f'I{destination_row}'] = "Team_D_CCH_PCC_1"
            elif market == 'France':
                ws[f'I{destination_row}'] = "Team_F_CCH_PCC_1"
            elif market == 'Spain':
                ws[f'I{destination_row}'] = "Team_E_CCH_PCC_1"
            elif market == 'Italy':
                ws[f'I{destination_row}'] = "Team_I_CCH_PCC_1"
        elif pcc_status == 'DS':
            if market == 'DACH':
                ws[f'I{destination_row}'] = "Team_D_CCH_B2C_1"
            elif market == 'France':
                ws[f'I{destination_row}'] = "Team_F_CCH_B2C_1"
            elif market == 'Spain':
                ws[f'I{destination_row}'] = "Team_E_CCH_B2C_1"

        # Rellenar columna M (Is PCC)
        if pcc_status == 'Y':
            ws[f'M{destination_row}'] = "Y"
        else:
            ws[f'M{destination_row}'] = "N"

        # Rellenar columnas Q, R y S (CCRM, CTI User y TTG UserID 1)
        ws[f'Q{destination_row}'] = email
        ws[f'R{destination_row}'] = email
        ws[f'S{destination_row}'] = b2e_username

        # Rellenar columna W (Campaign Level)
        if pcc_status == 'Y':
            ws[f'W{destination_row}'] = "Agent"
        elif pcc_status == 'N':
            ws[f'W{destination_row}'] = "Agent"
        elif pcc_status == 'TL':
            ws[f'W{destination_row}'] = "Team Leader"
        elif pcc_status == 'DS':
            ws[f'W{destination_row}'] = "Agent"

        # Incrementar destination_row para la siguiente inserción
        destination_row += 1

    # Formatear la fecha y hora
    now = datetime.now()
    formatted_date = now.strftime("%Y%m%d_%H%M")  # Formato YYYYMMDD_HHMM

    # Guardar el archivo actualizado con el nuevo nombre
    output_file = os.path.join(tempfile.gettempdir(), f'D365_STEP4_CCH_{formatted_date}.xlsx')
    wb.save(output_file)

    # Enviar el archivo descargable
    return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
